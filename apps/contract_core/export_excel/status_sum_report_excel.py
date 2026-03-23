# apps/contract_core/export_excel/status_sum_report_excel.py
from io import BytesIO
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from apps.contract_core.services.status_sum_report_service import StatusSumReportService
from apps.contract_core.models import Contract


class StatusSumReportExcelExporter:
    """
    Экспортер отчета по статусам и суммам в Excel.
    """

    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    SUBHEADER_FILL = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    SUBHEADER_FONT = Font(bold=True, size=10)
    TOTAL_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    TOTAL_FONT = Font(bold=True, size=10)
    CATEGORY_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    CATEGORY_FONT = Font(bold=True, size=10)

    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(self, user):
        self.user = user
        self.service = StatusSumReportService(user)
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Отчет по статусам"
        self.current_row = 1

    def _format_money(self, value: Decimal) -> str:
        """Форматирует сумму."""
        return f"{value:,.0f}".replace(",", " ")

    def _format_number(self, value: int) -> str:
        """Форматирует число."""
        return f"{value:,}".replace(",", " ")

    def _write_header(self):
        """Заголовок отчета."""
        date_str = self.service.report_date.strftime('%d.%m.%Y')
        title = f"Отчет по статусам и общим суммам неархивных договоров на {date_str}"

        self.ws.merge_cells(f'A{self.current_row}:E{self.current_row}')
        cell = self.ws.cell(self.current_row, 1, title)
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='center')
        self.current_row += 2

    def _write_company_header(self, company_data):
        """Заголовок компании."""
        roles_str = ", ".join(company_data.roles)
        title = f"{company_data.company.name} (ИНН {company_data.company.inn})\t\t\t\t{roles_str}"

        self.ws.merge_cells(f'A{self.current_row}:E{self.current_row}')
        cell = self.ws.cell(self.current_row, 1, title)
        cell.font = Font(bold=True, size=11)
        self.current_row += 1

    def _write_section_table(self, section):
        """Таблица по секции (типу контракта)."""
        # Заголовок категории
        self.ws.merge_cells(f'A{self.current_row}:E{self.current_row}')
        cell = self.ws.cell(self.current_row, 1, section['name'])
        cell.fill = self.CATEGORY_FILL
        cell.font = self.CATEGORY_FONT
        cell.border = self.BORDER
        self.current_row += 1

        # Заголовки колонок
        headers = ['', 'Всего', 'Завершен', 'Действует', 'Истекает']
        if not section['has_expired']:
            headers = ['', 'Всего', 'Завершен', 'Действует', 'Истекает']
            # Для долгосрочных убираем "Истекает" из заголовков или оставляем пустым
            headers[4] = ''  # Пустая колонка для долгосрочных

        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(self.current_row, col, header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER
        self.current_row += 1

        stats = section['stats']

        # Строка "Количество"
        row_data = [
            'Количество, шт.',
            self._format_number(stats['total']['count']),
            self._format_number(stats[Contract.STATUS_COMPLETED]['count']),
            self._format_number(stats[Contract.STATUS_ACTIVE]['count']),
        ]

        if section['has_expired']:
            row_data.append(self._format_number(stats[Contract.STATUS_ACTIVE_EXPIRES]['count']))
        else:
            row_data.append('')  # Пусто для долгосрочных

        for col, value in enumerate(row_data, 1):
            cell = self.ws.cell(self.current_row, col, value)
            cell.border = self.BORDER
            if col == 1:
                cell.font = self.TOTAL_FONT
            else:
                cell.alignment = Alignment(horizontal='center')
        self.current_row += 1

        # Строка "Сумма общая"
        row_data = [
            'Сумма общая, р.',
            self._format_money(stats['total']['sum']),
            self._format_money(stats[Contract.STATUS_COMPLETED]['sum']),
            self._format_money(stats[Contract.STATUS_ACTIVE]['sum']),
        ]

        if section['has_expired']:
            row_data.append(self._format_money(stats[Contract.STATUS_ACTIVE_EXPIRES]['sum']))
        else:
            row_data.append('')

        for col, value in enumerate(row_data, 1):
            cell = self.ws.cell(self.current_row, col, value)
            cell.border = self.BORDER
            if col == 1:
                cell.font = self.TOTAL_FONT
            else:
                cell.alignment = Alignment(horizontal='right')
        self.current_row += 2

    def _adjust_column_widths(self):
        """Настройка ширины колонок."""
        self.ws.column_dimensions['A'].width = 25
        self.ws.column_dimensions['B'].width = 15
        self.ws.column_dimensions['C'].width = 15
        self.ws.column_dimensions['D'].width = 15
        self.ws.column_dimensions['E'].width = 15

    def export(self) -> BytesIO:
        """Генерирует Excel-файл."""
        from apps.contract_core.models import Contract

        self._write_header()

        report_data = self.service.generate_report()

        for company_data in report_data['companies_data']:
            self._write_company_header(company_data)

            for section_key, section in company_data.sections.items():
                self._write_section_table(section)

            self.current_row += 1

        self._adjust_column_widths()

        output = BytesIO()
        self.wb.save(output)
        output.seek(0)

        return output

    def get_filename(self) -> str:
        """Имя файла."""
        date_str = self.service.report_date.strftime('%d.%m.%Y')
        return f"Отчет_по_статусам_и_суммам_{date_str}.xlsx"