# apps/contract_core/export_excel/works_sum_report_excel.py
from io import BytesIO
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from apps.contract_core.services.works_sum_report_service import WorksSumReportService


class WorksSumReportExcelExporter:
    """
    Экспортер отчета "Работы и суммы" в Excel.
    """

    # Стили
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    SUBHEADER_FILL = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    SUBHEADER_FONT = Font(bold=True, size=10)
    TOTAL_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    TOTAL_FONT = Font(bold=True, size=10)
    CATEGORY_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    CATEGORY_FONT = Font(bold=True, size=10)

    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(self, user):
        self.user = user
        self.service = WorksSumReportService(user)
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Отчет по работам"
        self.current_row = 1

    def _format_money(self, value: Decimal) -> str:
        """Форматирует сумму для Excel."""
        if value == 0:
            return ""
        return f"{value:,.0f}".replace(",", " ")

    def _format_number(self, value: int) -> str:
        """Форматирует число для Excel."""
        return f"{value:,}".replace(",", " ")

    def _write_header(self):
        """Записывает заголовок отчета."""
        report_date = self.service.report_date

        # Заголовок
        self.ws.merge_cells(f'A{self.current_row}:D{self.current_row}')
        cell = self.ws.cell(self.current_row, 1,
                            f"Отчет по работам и суммам действующих договоров на {report_date.strftime('%d.%m.%Y')}")
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='center')
        self.current_row += 2

    def _write_company_header(self, company_data):
        """Записывает заголовок компании."""
        # Название компании и роли
        roles_str = ", ".join(company_data.roles)
        company_title = f"{company_data.company.name} (ИНН {company_data.company.inn})\t{roles_str}\tдействующие договоры на {self.service.report_date.strftime('%d.%m.%Y')}"

        self.ws.merge_cells(f'A{self.current_row}:D{self.current_row}')
        cell = self.ws.cell(self.current_row, 1, company_title)
        cell.font = Font(bold=True, size=11)
        self.current_row += 1

    def _write_table_header(self):
        """Записывает заголовки таблицы."""
        headers = ["Вид работ / Категория", "Количество, шт.", "Сумма общая", "Сумма в месяц"]

        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(self.current_row, col, header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER

        self.current_row += 1

    def _write_total_row(self, label: str, count: int, total_sum: Decimal, monthly_sum: Decimal,
                         is_grand_total: bool = False):
        """Записывает строку итогов."""
        fill = self.TOTAL_FILL if is_grand_total else None
        font = self.TOTAL_FONT

        # Метка
        cell = self.ws.cell(self.current_row, 1, label)
        cell.font = font
        cell.border = self.BORDER
        if fill:
            cell.fill = fill

        # Количество
        cell = self.ws.cell(self.current_row, 2, self._format_number(count))
        cell.font = font
        cell.alignment = Alignment(horizontal='center')
        cell.border = self.BORDER
        if fill:
            cell.fill = fill

        # Сумма общая
        cell = self.ws.cell(self.current_row, 3, self._format_money(total_sum))
        cell.font = font
        cell.alignment = Alignment(horizontal='right')
        cell.border = self.BORDER
        if fill:
            cell.fill = fill

        # Сумма в месяц
        cell = self.ws.cell(self.current_row, 4, self._format_money(monthly_sum))
        cell.font = font
        cell.alignment = Alignment(horizontal='right')
        cell.border = self.BORDER
        if fill:
            cell.fill = fill

        self.current_row += 1

    def _write_category_header(self, name: str):
        """Записывает заголовок категории."""
        self.ws.merge_cells(f'A{self.current_row}:D{self.current_row}')
        cell = self.ws.cell(self.current_row, 1, name)
        cell.fill = self.CATEGORY_FILL
        cell.font = self.CATEGORY_FONT
        cell.border = self.BORDER
        self.current_row += 1

    def _write_work_row(self, work_name: str, stats: dict, indent: int = 0):
        """Записывает строку вида работ."""
        indent_str = "    " * indent

        # Название работы
        cell = self.ws.cell(self.current_row, 1, f"{indent_str}{work_name}")
        cell.border = self.BORDER

        # Количество
        cell = self.ws.cell(self.current_row, 2, self._format_number(stats['count']))
        cell.alignment = Alignment(horizontal='center')
        cell.border = self.BORDER

        # Сумма общая
        cell = self.ws.cell(self.current_row, 3, self._format_money(stats['sum']))
        cell.alignment = Alignment(horizontal='right')
        cell.border = self.BORDER

        # Сумма в месяц
        cell = self.ws.cell(self.current_row, 4, self._format_money(stats['monthly']))
        cell.alignment = Alignment(horizontal='right')
        cell.border = self.BORDER

        self.current_row += 1

    def _write_company_table(self, company_data):
        """Записывает таблицу по компании."""
        # Заголовки
        self._write_table_header()

        # Итого по компании
        self._write_total_row(
            "Всего",
            company_data.total_count,
            company_data.total_sum,
            company_data.total_monthly,
            is_grand_total=True
        )

        # Секции по типам работ
        for section_key, section in company_data.sections.items():
            # Заголовок категории
            self._write_category_header(section['name'])

            # Итого по категории
            self._write_total_row(
                f"{section['name']} (итого)",
                section['total_count'],
                section['total_sum'],
                section['total_monthly']
            )

            # Виды работ
            for work_name, stats in section['work_types'].items():
                self._write_work_row(work_name, stats, indent=1)

        # Пустая строка между компаниями
        self.current_row += 1

    def _adjust_column_widths(self):
        """Настраивает ширину колонок."""
        self.ws.column_dimensions['A'].width = 50
        self.ws.column_dimensions['B'].width = 18
        self.ws.column_dimensions['C'].width = 20
        self.ws.column_dimensions['D'].width = 20

    def export(self) -> BytesIO:
        """
        Генерирует Excel-файл и возвращает BytesIO.
        """
        # Заголовок отчета
        self._write_header()

        # Получаем данные
        report_data = self.service.generate_report()

        # Таблицы по компаниям
        for company_data in report_data['companies_data']:
            self._write_company_header(company_data)
            self._write_company_table(company_data)

        # Настройка ширины колонок
        self._adjust_column_widths()

        # Сохраняем в BytesIO
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)

        return output

    def get_filename(self) -> str:
        """Возвращает имя файла для скачивания."""
        date_str = self.service.report_date.strftime('%d.%m.%Y')
        return f"Отчет_по_работам_и_суммам_{date_str}.xlsx"