# apps/contract_core/export_excel/signing_stage_report_exporter.py
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class SigningStageReportExporter:
    """
    Сервис для экспорта отчёта по стадиям подписания в Excel.
    """

    def __init__(self, report_data, stages, control_days):
        self.report_data = report_data
        self.stages = stages
        self.control_days = control_days
        self.wb = None
        self.ws = None

    def export(self) -> BytesIO:
        """
        Генерирует Excel файл и возвращает BytesIO с данными.
        """
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Отчет по стадиям"

        self._setup_styles()
        self._fill_data()
        self._adjust_column_widths()
        self._freeze_panes()

        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output

    def _setup_styles(self):
        """Настройка стилей для ячеек."""
        self.styles = {
            'title': {
                'font': Font(bold=True, size=14),
                'alignment': Alignment(horizontal='center')
            },
            'company_header': {
                'font': Font(bold=True, size=12, color="FFFFFF"),
                'fill': PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            },
            'role_header': {
                'font': Font(bold=True, size=10),
                'fill': PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
            },
            'table_header': {
                'font': Font(color="FFFFFF", bold=True, size=11),
                'fill': PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid"),
                'alignment': Alignment(horizontal='center', wrap_text=True)
            },
            'total_bold': {
                'font': Font(bold=True),
                'alignment': Alignment(horizontal='center')
            },
            'overdue': {
                'font': Font(color="E74C3C", bold=True),  # Красный
                'alignment': Alignment(horizontal='center')
            },
            'note': {
                'font': Font(italic=True, size=9, color="7F8C8D")
            },
            'border': Border(
                left=Side(style='thin', color='BDC3C7'),
                right=Side(style='thin', color='BDC3C7'),
                top=Side(style='thin', color='BDC3C7'),
                bottom=Side(style='thin', color='BDC3C7')
            )
        }

    def _fill_data(self):
        """Заполнение данными."""
        row = 1

        # Заголовок отчета
        self._write_merged_cell(
            row, 1, f"A{row}:H{row}",
            f"Отчёт по стадиям подписания договоров (контрольный срок: {self.control_days} дн.)",
            self.styles['title']
        )
        row += 2

        for company_data in self.report_data:
            row = self._write_company_block(row, company_data)

        return row

    def _write_company_block(self, row, company_data):
        """Запись блока компании."""
        # Название компании
        self._write_merged_cell(
            row, 1, f"A{row}:H{row}",
            f"{company_data.company.name} (ИНН: {company_data.company.inn})",
            self.styles['company_header']
        )
        row += 1

        for role in company_data.roles:
            row = self._write_role_block(row, role)

        return row

    def _write_role_block(self, row, role):
        """Запись блока роли."""
        # Роль (Лицензиат/Лаборатория)
        self._write_merged_cell(
            row, 1, f"A{row}:H{row}",
            role.role_name,
            self.styles['role_header']
        )
        row += 1

        # Заголовки таблицы
        headers = ["Тип договора", "Всего"] + [s.name for s in self.stages]
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=row, column=col, value=header)
            self._apply_style(cell, self.styles['table_header'])
            cell.border = self.styles['border']
        row += 1

        # Данные по типам договоров
        for type_data in role.contract_types:
            row = self._write_contract_type_row(row, type_data)

        # Примечание
        note = (f"Примечание: в скобках указано количество контрактов с превышением "
                f"контрольного срока ({self.control_days} дн.)")
        self._write_merged_cell(
            row, 1, f"A{row}:H{row}",
            note,
            self.styles['note']
        )
        row += 2

        return row

    def _write_contract_type_row(self, row, type_data):
        """Запись строки с данными по типу договора."""
        col = 1

        # Тип договора
        cell = self.ws.cell(row=row, column=col, value=type_data.type_label)
        cell.border = self.styles['border']
        cell.alignment = Alignment(horizontal='left')
        col += 1

        # Всего
        cell = self.ws.cell(row=row, column=col, value=type_data.total)
        cell.border = self.styles['border']
        self._apply_style(cell, self.styles['total_bold'])
        col += 1

        # Стадии
        for stage_count in type_data.stage_counts:
            self._write_stage_cell(row, col, stage_count)
            col += 1

        return row + 1

    def _write_stage_cell(self, row, col, stage_count):
        """Запись ячейки со стадией."""
        if stage_count.count > 0:
            if stage_count.overdue_count > 0:
                value = f"{stage_count.count} ({stage_count.overdue_count})"
                cell = self.ws.cell(row=row, column=col, value=value)
                self._apply_style(cell, self.styles['overdue'])
            else:
                cell = self.ws.cell(row=row, column=col, value=stage_count.count)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
        else:
            cell = self.ws.cell(row=row, column=col, value="")

        cell.border = self.styles['border']
        cell.alignment = Alignment(horizontal='center')

    def _write_merged_cell(self, row, col, range_str, value, style_dict):
        """Запись объединенной ячейки."""
        self.ws.merge_cells(range_str)
        cell = self.ws.cell(row=row, column=col, value=value)
        self._apply_style(cell, style_dict)

    def _apply_style(self, cell, style_dict):
        """Применение стиля к ячейке."""
        if 'font' in style_dict:
            cell.font = style_dict['font']
        if 'fill' in style_dict:
            cell.fill = style_dict['fill']
        if 'alignment' in style_dict:
            cell.alignment = style_dict['alignment']

    def _adjust_column_widths(self):
        """Настройка ширины колонок."""
        # Первая колонка шире (тип договора)
        self.ws.column_dimensions['A'].width = 30

        # Остальные колонки
        for col in range(2, len(self.stages) + 3):
            self.ws.column_dimensions[get_column_letter(col)].width = 18

    def _freeze_panes(self):
        """Заморозка заголовков."""
        self.ws.freeze_panes = 'A3'