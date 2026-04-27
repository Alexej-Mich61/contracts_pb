# apps/contract_core/export_excel/contract_detail_excel.py
from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def export_contract_detail_to_excel(contract, objects_count, aks_count, filename="contract_detail.xlsx"):
    """
    Формирует Excel с детальной информацией по договору.
    Структура повторяет секции модального окна.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Детали договора"

    # ── Стили ──
    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    section_font = Font(bold=True, size=11, color="FFFFFF")
    section_align = Alignment(horizontal="left", vertical="center")

    label_font = Font(bold=True, size=10)
    value_font = Font(size=10)
    cell_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    fills = {
        "primary": PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid"),
        "success": PatternFill(start_color="198754", end_color="198754", fill_type="solid"),
        "info":    PatternFill(start_color="0dcaf0", end_color="0dcaf0", fill_type="solid"),
        "warning": PatternFill(start_color="ffc107", end_color="ffc107", fill_type="solid"),
        "danger":  PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid"),
        "light":   PatternFill(start_color="e9ecef", end_color="e9ecef", fill_type="solid"),
    }

    # ── Хелперы ──
    def section_header(row, title, color="primary", merge_to=6):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_to)
        for col in range(1, merge_to + 1):
            c = ws.cell(row=row, column=col)
            c.border = thin_border
            c.fill = fills[color]
        c = ws.cell(row=row, column=1, value=title)
        c.font = section_font
        c.alignment = section_align
        return row + 1

    def kv_row(row, key, value, merge_to=6):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=merge_to)
        c1 = ws.cell(row=row, column=1, value=key)
        c1.font = label_font
        c1.border = thin_border
        c1.alignment = cell_align

        c2 = ws.cell(row=row, column=2, value=value)
        c2.font = value_font
        c2.border = thin_border
        c2.alignment = cell_align
        return row + 1

    # ── Ширина колонок ──
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 35

    r = 1

    # ═══════════════════════════════════════
    # ЗАГОЛОВОК (как шапка модалки)
    # ═══════════════════════════════════════
    signing = (
        contract.signing_stage.stage.name
        if hasattr(contract, "signing_stage") and contract.signing_stage
        else "Стадия не указана"
    )
    header_text = f"Договор №{contract.number or 'б/н'} — {contract.get_type_display()} — {signing}"

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    for col in range(1, 7):
        c = ws.cell(row=r, column=col)
        c.border = thin_border
        c.fill = fills["primary"]
    c = ws.cell(row=r, column=1, value=header_text)
    c.font = title_font
    c.alignment = title_align
    ws.row_dimensions[r].height = 30
    r += 2

    # ═══════════════════════════════════════
    # 1. ОСНОВНЫЕ СВЕДЕНИЯ  (border-primary)
    # ═══════════════════════════════════════
    r = section_header(r, "Основные сведения", "primary")
    r = kv_row(r, "Тип договора:", contract.get_type_display())
    r = kv_row(r, "ID:", contract.pk)
    r = kv_row(r, "Заказчик:", f"{contract.customer.name} (ИНН: {contract.customer.inn})" if contract.customer else "—")
    r = kv_row(r, "Номер и дата:", f"№ {contract.number or 'б/н'} от {contract.date_concluded.strftime('%d.%m.%Y') if contract.date_concluded else '—'}")
    r = kv_row(r, "Исполнитель:", f"{contract.executor.name} (ИНН: {contract.executor.inn})" if contract.executor else "—")
    r = kv_row(r, "Срок:", f"{contract.date_start.strftime('%d.%m.%Y') if contract.date_start else '—'} — {contract.date_end.strftime('%d.%m.%Y') if contract.date_end else '—'}")
    r = kv_row(r, "Статус:", contract.get_status_display())
    r = kv_row(r, "Работы:", contract.work.name if contract.work else "—")
    r += 1

    # ═══════════════════════════════════════
    # 2. ФИНАНСЫ  (border-success)
    # ═══════════════════════════════════════
    r = section_header(r, "Финансы", "success")
    r = kv_row(r, "Сумма общая:", f"{contract.total_sum or 0:,.2f} руб.")
    r = kv_row(r, "Сумма в месяц:", f"{contract.monthly_sum or 0:,.2f} руб.")
    r = kv_row(r, "Аванс:", f"{contract.advance or 0:,.2f} руб.")
    r = kv_row(r, "Файл договора:", "Скачать" if contract.file else "—")
    r += 1

    # ═══════════════════════════════════════
    # 3. ОБЪЕКТЫ ЗАЩИТЫ  (border-info)
    # ═══════════════════════════════════════
    r = section_header(r, f"Объекты защиты (Объектов: {objects_count}, АК: {aks_count})", "info")

    contract_objects = list(contract.contract_objects.all())
    if contract_objects:
        # заголовки таблицы
        headers = ["Объект", "Контакты", "Район / Регион", "Адрес", "Субподрядчик", "АК"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=r, column=col, value=h)
            c.font = label_font
            c.border = thin_border
            c.fill = fills["light"]
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        r += 1

        for obj in contract_objects:
            aks_lines = []
            for ak in obj.aks.all():
                aks_lines.append(f"АК {ak.number} — {ak.name} ({ak.address or 'адрес не указан'})")
            aks_str = "\n".join(aks_lines) if aks_lines else "—"

            region = obj.district.region.name if obj.district and obj.district.region else "—"
            district = obj.district.name if obj.district else "—"

            subcontractor_text = "—"
            if obj.subcontractor:
                subcontractor_text = (
                    f"{obj.subcontractor.name}\n"
                    f"общ.: {obj.total_sum_subcontract or 0:,.2f} руб., "
                    f"мес.: {obj.monthly_sum_subcontract or 0:,.2f} руб."
                )

            vals = [
                obj.name,
                obj.contacts or "—",
                f"{district} ({region})",
                obj.address or "—",
                subcontractor_text,
                aks_str,
            ]

            max_lines = max(len(str(v).split("\n")) for v in vals)
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=val)
                c.border = thin_border
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                c.font = value_font

            ws.row_dimensions[r].height = max(15 * max_lines, 28)
            r += 1
    else:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1, value="Нет объектов защиты")
        c.alignment = cell_align
        r += 1
    r += 1

    # ═══════════════════════════════════════
    # 4. ПОДПИСАНИЕ  (border-warning)
    # ═══════════════════════════════════════
    r = section_header(r, "Подписание", "warning")
    if contract.signing_stage:
        r = kv_row(r, "Стадия:", contract.signing_stage.stage.name)
        changed = contract.signing_stage.changed_at.strftime("%d.%m.%Y %H:%M") if contract.signing_stage.changed_at else "—"
        by = contract.signing_stage.changed_by.get_full_name() if contract.signing_stage.changed_by else "—"
        r = kv_row(r, "Изменено:", f"{changed} — {by}")
        if contract.signing_stage.note:
            r = kv_row(r, "Примечание:", contract.signing_stage.note)
    else:
        r = kv_row(r, "Стадия:", "Не указана")
    r += 1

    # ═══════════════════════════════════════
    # 5. СИСТЕМЫ  (border-warning)
    # ═══════════════════════════════════════
    r = section_header(r, "Системы", "warning")
    sys_checks = list(contract.system_checks.all())
    if sys_checks:
        headers = ["Система", "Дата проверки", "Кто отметил", "Примечание"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=r, column=col, value=h)
            c.font = label_font
            c.border = thin_border
            c.fill = fills["light"]
            c.alignment = Alignment(horizontal="center", vertical="center")
        r += 1

        for check in sys_checks:
            vals = [
                check.system_type.name,
                check.last_checked.strftime("%d.%m.%Y") if check.last_checked else "не отмечено",
                check.checked_by.get_full_name() if check.checked_by else "—",
                check.note or "—",
            ]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=val)
                c.border = thin_border
                c.alignment = cell_align
                c.font = value_font
            r += 1
    else:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(row=r, column=1, value="Нет отметок")
        c.alignment = cell_align
        r += 1
    r += 1

    # ═══════════════════════════════════════
    # 6. ИТОГОВЫЙ АКТ  (border-danger)
    # ═══════════════════════════════════════
    r = section_header(r, "Итоговый акт", "danger")
    if contract.final_act:
        r = kv_row(r, "Наличие:", "Да" if contract.final_act.present else "Нет")
        r = kv_row(r, "Дата:", contract.final_act.date.strftime("%d.%m.%Y") if contract.final_act.date else "—")
        r = kv_row(r, "Файл:", "Есть" if contract.final_act.file else "—")
    else:
        r = kv_row(r, "Наличие:", "Нет итогового акта")
    r += 1

    # ═══════════════════════════════════════
    # 7. ПРОМЕЖУТОЧНЫЕ АКТЫ  (border-danger)
    # ═══════════════════════════════════════
    r = section_header(r, f"Промежуточные акты ({contract.interim_acts.count()})", "danger")
    interim = list(contract.interim_acts.all())
    if interim:
        headers = ["Название", "Дата", "Файл"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=r, column=col, value=h)
            c.font = label_font
            c.border = thin_border
            c.fill = fills["light"]
            c.alignment = Alignment(horizontal="center", vertical="center")
        r += 1

        for act in interim:
            vals = [
                act.title,
                act.date.strftime("%d.%m.%Y") if act.date else "—",
                "Есть" if act.file else "—",
            ]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=val)
                c.border = thin_border
                c.alignment = cell_align
                c.font = value_font
            r += 1
    else:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        c = ws.cell(row=r, column=1, value="Нет промежуточных актов")
        c.alignment = cell_align
        r += 1
    r += 1

    # ═══════════════════════════════════════
    # 8. СЛУЖЕБНОЕ  (border-info)
    # ═══════════════════════════════════════
    r = section_header(r, "Служебное", "info")
    r = kv_row(r, "Создал:", contract.creator.get_full_name() if contract.creator else "—")
    r = kv_row(r, "Обновил:", contract.updater.get_full_name() if contract.updater else "—")
    r = kv_row(r, "В корзине:", "Да" if contract.is_trash else "Нет")
    r = kv_row(r, "В архиве:", "Да" if contract.is_archived else "Нет")

    # ── Ответ ──
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response