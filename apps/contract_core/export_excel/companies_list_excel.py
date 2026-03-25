# apps/contract_core/export_excel/companies_list_excel.py
"""
Модуль для экспорта списка компаний в Excel.
Содержит логику формирования Excel-файла, отделённую от вьюхи.
"""

from io import BytesIO
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from apps.contract_core.models import Company


class CompaniesExcelExporter:
    """
    Класс для экспорта списка компаний в Excel.

    Ответственность: формирование Excel-файла с заданными колонками
    и стилями, оптимизированными для печати на листе А4.
    """

    # Порядок и настройки колонок
    COLUMNS = [
        {'key': 'id', 'header': 'ID', 'width': 8},
        {'key': 'name', 'header': 'Название', 'width': 35},
        {'key': 'inn', 'header': 'ИНН', 'width': 15},
        {'key': 'email', 'header': 'E-mail', 'width': 25},
        {'key': 'phone', 'header': 'Телефон', 'width': 15},
        {'key': 'roles', 'header': 'Роли', 'width': 25},
        {'key': 'description', 'header': 'Описание', 'width': 40},
    ]

    # Стили
    HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

    CELL_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    CELL_BORDER = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    def __init__(self, companies: List[Company]):
        """
        Args:
            companies: QuerySet или список объектов Company для экспорта
        """
        self.companies = companies

    def _get_roles_display(self, company: Company) -> str:
        """Формирует строку с ролями компании."""
        roles = []
        if company.is_customer:
            roles.append("Заказчик")
        if company.is_licensee:
            roles.append("Лицензиат МЧС")
        if company.is_laboratory:
            roles.append("Лаборатория")
        if company.is_subcontractor:
            roles.append("Субподрядчик")
        return ", ".join(roles) if roles else "—"

    def _prepare_row_data(self, company: Company) -> dict:
        """Подготавливает данные строки для Excel."""
        return {
            'id': company.id,
            'name': company.name or "—",
            'inn': company.inn or "—",
            'email': company.email or "—",
            'phone': company.phone or "—",
            'roles': self._get_roles_display(company),
            'description': company.description or "—",
        }

    def _setup_worksheet(self, ws) -> None:
        """Настройка листа Excel для печати на А4."""
        # Настройки страницы для печати на А4
        ws.page_setup.paperSize = 9  # A4
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        # Поля
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75

        # Заголовок и колонтитулы
        ws.oddHeader.center.text = "Справочник организаций"
        ws.oddHeader.center.size = 12
        ws.oddHeader.center.font = "Arial,Bold"

        ws.oddFooter.center.text = "Страница &P из &N"
        ws.oddFooter.center.size = 10

    def _apply_header_style(self, cell) -> None:
        """Применяет стиль к ячейке заголовка."""
        cell.font = self.HEADER_FONT
        cell.fill = self.HEADER_FILL
        cell.alignment = self.HEADER_ALIGNMENT
        cell.border = self.CELL_BORDER

    def _apply_cell_style(self, cell) -> None:
        """Применяет стиль к обычной ячейке."""
        cell.alignment = self.CELL_ALIGNMENT
        cell.border = self.CELL_BORDER

    def export(self, filename: Optional[str] = None) -> BytesIO:
        """
        Создаёт Excel-файл и возвращает его как BytesIO.

        Args:
            filename: Имя файла (опционально, для метаданных)

        Returns:
            BytesIO с содержимым Excel-файла
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Организации"

        # Настройка листа
        self._setup_worksheet(ws)

        # Заголовки колонок
        for col_idx, col_config in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_config['header'])
            self._apply_header_style(cell)
            # Установка ширины колонки
            ws.column_dimensions[get_column_letter(col_idx)].width = col_config['width']

        # Фиксация шапки
        ws.freeze_panes = 'A2'

        # Автофильтр
        ws.auto_filter.ref = f"A1:{get_column_letter(len(self.COLUMNS))}1"

        # Данные
        for row_idx, company in enumerate(self.companies, 2):
            row_data = self._prepare_row_data(company)
            for col_idx, col_config in enumerate(self.COLUMNS, 1):
                value = row_data.get(col_config['key'], '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                self._apply_cell_style(cell)

        # Фиксированная высота строк заголовка
        ws.row_dimensions[1].height = 30

        # Сохранение в BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    def get_filename(self) -> str:
        """Генерирует имя файла для скачивания."""
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"companies_export_{timestamp}.xlsx"


def export_companies_to_excel(companies: List[Company]) -> tuple[BytesIO, str]:
    """
    Утилитарная функция для быстрого экспорта.

    Args:
        companies: QuerySet или список объектов Company

    Returns:
        Кортеж (BytesIO с файлом, имя файла)
    """
    exporter = CompaniesExcelExporter(companies)
    file_obj = exporter.export()
    filename = exporter.get_filename()
    return file_obj, filename