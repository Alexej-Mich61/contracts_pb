# apps/contract_core/export_excel/contract_list_excel.py
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def export_contracts_to_excel(contracts, filename="contracts.xlsx"):
    """
    Формирует Excel-файл со списком договоров.
    Принимает queryset (уже отфильтрованный).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Договоры"

    headers = [
        "№ договора и дата",
        "Заказчик",
        "ИНН Заказчика",
        "Исполнитель",
        "Срок действия",
        "Сумма общая, ₽",
        "Сумма в месяц, ₽",
        "Аванс, ₽",
        "Кол-во объектов",
        "Кол-во АК",
        "Работы",
        "Стадия подписания",
        "Статус",
        "Оплата",
        "Итоговый акт",
    ]

    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Заголовки
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Данные
    for row_num, contract in enumerate(contracts, 2):
        number_date = (
            f"{contract.number or 'б/н'} от "
            f"{contract.date_concluded.strftime('%d.%m.%Y') if contract.date_concluded else '—'}"
        )
        period = (
            f"{contract.date_start.strftime('%d.%m.%Y') if contract.date_start else '—'} — "
            f"{contract.date_end.strftime('%d.%m.%Y') if contract.date_end else '—'}"
        )
        signing = (
            contract.signing_stage.stage.name
            if hasattr(contract, "signing_stage") and contract.signing_stage
            else "—"
        )

        # Подсчёт объектов и АК (если нет аннотаций во вьюхе)
        obj_count = getattr(contract, "object_count", None)
        if obj_count is None:
            obj_count = contract.contract_objects.count()

        ak_count = getattr(contract, "ak_count", None)
        if ak_count is None:
            ak_count = sum(obj.aks.count() for obj in contract.contract_objects.all())

        # --- НОВЫЕ ПОЛЯ ---
        customer_inn = contract.customer.inn if contract.customer else "—"
        payment = "Да" if contract.is_paid else ""
        final_act_status = ""
        if hasattr(contract, "final_act") and contract.final_act and contract.final_act.present:
            final_act_status = "Есть"

        values = [
            number_date,
            contract.customer.name if contract.customer else "—",
            customer_inn,
            contract.executor.name if contract.executor else "—",
            period,
            contract.total_sum or 0,
            contract.monthly_sum or 0,
            contract.advance or 0,
            obj_count,
            ak_count,
            contract.work.name if contract.work else "—",
            signing,
            contract.get_status_display(),
            payment,
            final_act_status,
        ]

        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border

            if col_num in (6, 7, 8):          # деньги — выравнивание вправо + числовой формат
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_num in (9, 10):           # количества — по центру
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:                             # текст — влево + перенос
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Ширина колонок
    column_widths = [22, 30, 16, 30, 24, 16, 16, 14, 14, 12, 28, 22, 16, 12, 16]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Закрепить заголовок
    ws.freeze_panes = "A2"

    # Автофильтр (охватывает все колонки автоматически)
    ws.auto_filter.ref = ws.dimensions

    # Ответ
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response